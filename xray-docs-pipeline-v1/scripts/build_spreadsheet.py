#!/usr/bin/env python3
"""
build_spreadsheet.py — Financial Model Generator

Generates a structured financial spreadsheet (.xlsx) from a Business Case document.

Usage:
    python scripts/build_spreadsheet.py \
        --business-case ./output/corpus/business-case.md \
        --project-name "My Project" \
        --output-dir ./output/final

Dependencies:
    pip install openpyxl
"""

import argparse
import re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("Warning: openpyxl not installed. Run: pip install openpyxl")


# ─── Style constants ──────────────────────────────────────────────────────────

COLORS = {
    "primary":      "1A1A2E",
    "accent":       "E94560",
    "header_bg":    "16213E",
    "subheader_bg": "0F3460",
    "positive":     "27AE60",
    "negative":     "E74C3C",
    "neutral":      "F39C12",
    "light_bg":     "F8F9FA",
    "white":        "FFFFFF",
    "border":       "DEE2E6",
}


def make_font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def make_border(style="thin"):
    s = Side(style=style, color=COLORS["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header(ws, row, col, value, level="primary"):
    cell = ws.cell(row=row, column=col, value=value)
    if level == "primary":
        cell.font = make_font(bold=True, size=12, color=COLORS["white"])
        cell.fill = make_fill(COLORS["header_bg"])
    elif level == "secondary":
        cell.font = make_font(bold=True, size=11, color=COLORS["white"])
        cell.fill = make_fill(COLORS["subheader_bg"])
    elif level == "section":
        cell.font = make_font(bold=True, size=10, color=COLORS["header_bg"])
        cell.fill = make_fill("E8EAF6")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()
    return cell


def apply_data_cell(ws, row, col, value, fmt=None, bold=False, color=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = make_font(bold=bold, color=color or "000000")
    cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
    cell.border = make_border()
    if fmt:
        cell.number_format = fmt
    return cell


# ─── Sheet builders ───────────────────────────────────────────────────────────

def build_summary_sheet(wb, project_name: str, profile: str):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"{project_name} — Financial Model"
    title_cell.font = make_font(bold=True, size=16, color=COLORS["white"])
    title_cell.fill = make_fill(COLORS["primary"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = f"Profile: {profile}  |  Generated: {datetime.now().strftime('%Y-%m-%d')}"
    sub.font = make_font(size=10, color="888888", italic=True)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # KPI headers row
    kpi_labels = ["MRR Y1", "MRR Y3", "CAC", "LTV", "LTV/CAC", "Break-even", "Margin Y3"]
    kpi_defaults = ["R$ —", "R$ —", "R$ —", "R$ —", "—", "— months", "—%"]

    ws.row_dimensions[4].height = 24
    for i, (label, val) in enumerate(zip(kpi_labels, kpi_defaults)):
        apply_header(ws, 4, i + 1, label, level="secondary")
        apply_data_cell(ws, 5, i + 1, val)

    # Instructions
    ws.merge_cells("A7:G9")
    inst = ws["A7"]
    inst.value = (
        "This model is pre-structured. Fill in the 'Assumptions' sheet first, "
        "then review 'P&L', 'Unit Economics', and 'Scenarios'. "
        "All cells with yellow background are inputs. All others are calculated."
    )
    inst.font = make_font(size=10, italic=True, color="555555")
    inst.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    for col, width in enumerate([18, 18, 18, 18, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_assumptions_sheet(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False

    apply_header(ws, 1, 1, "Model Assumptions", level="primary")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")  # light yellow = input cell

    sections = [
        ("Revenue", [
            ("Avg. monthly price (R$)", 300, "BRL"),
            ("New customers / month (Year 1)", 10, "count"),
            ("New customers / month (Year 2)", 20, "count"),
            ("New customers / month (Year 3)", 35, "count"),
            ("Monthly churn rate (%)", 0.05, "pct"),
            ("Upsell rate (%)", 0.10, "pct"),
            ("Upsell avg. increase (R$)", 100, "BRL"),
        ]),
        ("Costs", [
            ("COGS per customer/month (R$)", 50, "BRL"),
            ("Salaries / month (R$)", 25000, "BRL"),
            ("Marketing spend / month (R$)", 5000, "BRL"),
            ("Infrastructure / month (R$)", 3000, "BRL"),
            ("Other fixed costs / month (R$)", 2000, "BRL"),
        ]),
        ("Unit Economics", [
            ("CAC (R$)", 800, "BRL"),
            ("Avg. customer lifetime (months)", 24, "count"),
            ("Support cost / customer / month (R$)", 30, "BRL"),
        ]),
        ("Funding", [
            ("Initial investment / runway (R$)", 500000, "BRL"),
            ("Months of runway sought", 18, "count"),
        ]),
    ]

    row = 3
    for section_name, items in sections:
        # Section header
        apply_header(ws, row, 1, section_name, level="section")
        ws.merge_cells(f"A{row}:D{row}")
        ws.row_dimensions[row].height = 22
        row += 1

        # Column headers
        for col, header in enumerate(["Parameter", "Value", "Unit", "Notes"], 1):
            apply_header(ws, row, col, header, level="secondary")
        ws.row_dimensions[row].height = 20
        row += 1

        for param, default, unit in items:
            ws.cell(row=row, column=1, value=param).font = make_font(size=10)
            val_cell = ws.cell(row=row, column=2, value=default)
            val_cell.fill = INPUT_FILL
            val_cell.font = make_font(bold=True, size=10)
            val_cell.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=3, value=unit).font = make_font(size=10, italic=True, color="888888")
            ws.cell(row=row, column=4, value="[HIPÓTESE] — update with real data").font = make_font(size=9, italic=True, color="AAAAAA")
            row += 1

        row += 1

    for col, width in enumerate([35, 15, 12, 40], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_pnl_sheet(wb):
    ws = wb.create_sheet("P&L (3 Years)")
    ws.sheet_view.showGridLines = False

    apply_header(ws, 1, 1, "Profit & Loss — 3 Year Projection", level="primary")
    ws.merge_cells("A1:N1")
    ws.row_dimensions[1].height = 28

    months = [f"M{i}" for i in range(1, 37)]
    year_labels = ["Year 1 (M1–M12)", "Year 2 (M13–M24)", "Year 3 (M25–M36)"]

    # Year group headers
    apply_header(ws, 2, 2, year_labels[0], level="secondary")
    ws.merge_cells("B2:M2")
    apply_header(ws, 2, 14, year_labels[1], level="secondary")
    ws.merge_cells("N2:Y2")
    apply_header(ws, 2, 26, year_labels[2], level="secondary")
    ws.merge_cells("Z2:AK2")

    # Month headers (abbreviated)
    apply_header(ws, 3, 1, "Line Item", level="secondary")
    for i, m in enumerate(months):
        apply_header(ws, 3, i + 2, m, level="secondary")
    ws.column_dimensions["A"].width = 28

    # P&L rows (placeholders — formulas should reference Assumptions sheet)
    pnl_rows = [
        ("REVENUE", True),
        ("  New MRR", False),
        ("  Expansion MRR", False),
        ("  Churned MRR", False),
        ("  Total MRR", False),
        ("", True),
        ("COST OF REVENUE", True),
        ("  COGS", False),
        ("  Support Costs", False),
        ("  Gross Profit", False),
        ("", True),
        ("OPERATING EXPENSES", True),
        ("  Salaries", False),
        ("  Marketing", False),
        ("  Infrastructure", False),
        ("  Other Fixed", False),
        ("  Total OpEx", False),
        ("", True),
        ("EBITDA", True),
        ("Net Profit / Loss", True),
        ("Cumulative Cash", True),
    ]

    for i, (label, is_section) in enumerate(pnl_rows):
        row = i + 4
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = make_font(bold=is_section, size=10)
        if is_section and label:
            cell.fill = make_fill("F0F0F0")
        for col in range(2, 38):
            data_cell = ws.cell(row=row, column=col, value=0 if label else "")
            data_cell.number_format = '#,##0'
            data_cell.font = make_font(size=9)
            data_cell.alignment = Alignment(horizontal="right")

    ws.row_dimensions[3].height = 20

    note = ws.cell(row=len(pnl_rows) + 6, column=1,
                   value="⚠ Fill in Assumptions sheet and replace placeholder values with formulas referencing those cells.")
    note.font = make_font(size=9, italic=True, color="E74C3C")


def build_unit_economics_sheet(wb):
    ws = wb.create_sheet("Unit Economics")
    ws.sheet_view.showGridLines = False

    apply_header(ws, 1, 1, "Unit Economics", level="primary")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 28

    metrics = [
        ("CAC (Customer Acquisition Cost)", "= Assumptions!B14", "R$"),
        ("ARPU (Avg Revenue Per User)", "= Assumptions!B2", "R$/month"),
        ("Gross Margin per Customer", "= ARPU - COGS - Support", "%"),
        ("Customer Lifetime (months)", "= Assumptions!B15", "months"),
        ("LTV (Lifetime Value)", "= ARPU × Lifetime × Gross Margin", "R$"),
        ("LTV / CAC Ratio", "= LTV / CAC", "ratio"),
        ("CAC Payback Period", "= CAC / (ARPU × Gross Margin)", "months"),
        ("Monthly Churn Rate", "= Assumptions!B5", "%"),
        ("Annual Churn Rate", "= 1 - (1 - Monthly Churn)^12", "%"),
    ]

    apply_header(ws, 3, 1, "Metric", level="secondary")
    apply_header(ws, 3, 2, "Formula / Value", level="secondary")
    apply_header(ws, 3, 3, "Unit", level="secondary")

    for i, (metric, formula, unit) in enumerate(metrics):
        row = i + 4
        ws.cell(row=row, column=1, value=metric).font = make_font(size=10)
        val = ws.cell(row=row, column=2, value=formula)
        val.font = make_font(bold=True, size=10)
        val.fill = PatternFill("solid", fgColor="FFF9C4")
        ws.cell(row=row, column=3, value=unit).font = make_font(size=10, italic=True, color="888888")

    for col, width in enumerate([38, 30, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_scenarios_sheet(wb):
    ws = wb.create_sheet("Scenarios")
    ws.sheet_view.showGridLines = False

    apply_header(ws, 1, 1, "Scenario Analysis", level="primary")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    headers = ["Metric", "Conservative", "Base Case", "Optimistic", "Label", "Notes"]
    for i, h in enumerate(headers):
        apply_header(ws, 3, i + 1, h, level="secondary")

    scenarios = [
        ("Monthly new customers (Y1)", 5, 10, 20, "[HIPÓTESE]", ""),
        ("Monthly churn rate", "8%", "5%", "3%", "[HIPÓTESE]", ""),
        ("Avg. price (R$)", 200, 300, 400, "[HIPÓTESE]", "Market test needed"),
        ("CAC (R$)", 1200, 800, 500, "[HIPÓTESE]", ""),
        ("Break-even (months)", 24, 18, 12, "[HIPÓTESE]", ""),
        ("MRR at 12 months", "R$ 10k", "R$ 30k", "R$ 60k", "[HIPÓTESE]", ""),
        ("MRR at 36 months", "R$ 50k", "R$ 150k", "R$ 350k", "[HIPÓTESE]", ""),
        ("Gross Margin Y3", "35%", "55%", "70%", "[HIPÓTESE]", ""),
    ]

    SCENARIO_FILLS = {
        1: PatternFill("solid", fgColor="FFEBEE"),  # Conservative — red tint
        2: PatternFill("solid", fgColor="E8F5E9"),  # Base — green tint
        3: PatternFill("solid", fgColor="E3F2FD"),  # Optimistic — blue tint
    }

    for i, row_data in enumerate(scenarios):
        row = i + 4
        ws.cell(row=row, column=1, value=row_data[0]).font = make_font(size=10)
        for col in range(1, 4):
            val_cell = ws.cell(row=row, column=col + 1, value=row_data[col])
            val_cell.fill = SCENARIO_FILLS[col]
            val_cell.font = make_font(bold=(col == 2), size=10)
            val_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=row_data[4]).font = make_font(size=9, italic=True, color="888888")
        ws.cell(row=row, column=6, value=row_data[5]).font = make_font(size=9, italic=True, color="AAAAAA")

    for col, width in enumerate([35, 18, 18, 18, 14, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate financial spreadsheet from business case")
    parser.add_argument("--business-case", help="Path to approved business-case.md")
    parser.add_argument("--project-name", required=True, help="Project name")
    parser.add_argument("--profile", default="product", help="Workflow profile")
    parser.add_argument("--output-dir", default="./output/final", help="Output directory")
    args = parser.parse_args()

    if not XLSX_AVAILABLE:
        print("openpyxl not installed. Run: pip install openpyxl")
        return

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    build_summary_sheet(wb, args.project_name, args.profile)
    build_assumptions_sheet(wb)
    build_pnl_sheet(wb)
    build_unit_economics_sheet(wb)
    build_scenarios_sheet(wb)

    slug = re.sub(r"[^a-z0-9]+", "-", args.project_name.lower()).strip("-")
    date_str = datetime.now().strftime("%Y%m%d")
    output_path = output_dir / f"{slug}-financial-model-{date_str}.xlsx"

    wb.save(str(output_path))
    print(f"✓ Financial model saved: {output_path}")


if __name__ == "__main__":
    main()
